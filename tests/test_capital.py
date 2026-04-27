import io

import pytest
from openpyxl import Workbook

from scrapers.capital import (
    parse_capital_holdings,
    parse_capital_xlsx,
    parse_capital_meta_ssr,
    find_latest_manual_xlsx,
)


@pytest.mark.parametrize("ticker, top_id, top_name", [
    ("00982A", "2330", "台積電"),
    ("00992A", "2330", "台積電"),
])
def test_parse_capital_ssr(ticker, top_id, top_name, fixture_path):
    text = fixture_path(f"capital_{ticker}.html").read_text(encoding="utf-8")
    holdings = parse_capital_holdings(text)

    # SSR exposes only top-10 — see docs/known-issues/capital-scraper.md
    assert 1 <= len(holdings) <= 10, f"expected 1-10 holdings, got {len(holdings)}"

    top = holdings[0]
    assert top.stock_id == top_id
    assert top.stock_name == top_name
    assert top.weight_pct > 0

    for h in holdings:
        assert h.stock_id
        assert h.stock_name
        assert h.weight_pct > 0
        assert h.shares > 0


# --- xlsx parser (manual upload path) ---

def _build_synthetic_xlsx(stocks: list[tuple[str, str, float, int]]) -> bytes:
    """Construct an xlsx matching Capital's "下載資料" structure for testing."""
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)
    ws = wb.create_sheet("股票")
    ws.append(["股票代號", "股票名稱", "持股權重(%)", "股數"])
    for stock_id, name, weight, shares in stocks:
        ws.append([stock_id, name, f"{weight}%", f"{shares:,}"])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_parse_capital_xlsx_full_holdings():
    """Full ~50-stock xlsx parses cleanly into TW-classified Holdings."""
    stocks = [
        ("2330", "台積電", 8.71, 1_548_000),
        ("5536", "聖暉*", 8.39, 3_355_000),
        ("3105", "穩懋", 5.56, 3_946_000),
    ]
    holdings = parse_capital_xlsx(_build_synthetic_xlsx(stocks))
    assert len(holdings) == 3
    assert holdings[0].stock_id == "2330"
    assert holdings[0].stock_name == "台積電"
    assert holdings[0].weight_pct == 8.71
    assert holdings[0].shares == 1_548_000
    assert holdings[0].market == "TW"
    # asterisk in name preserved (data-as-source-stores-it)
    assert holdings[1].stock_name == "聖暉*"


def test_parse_capital_xlsx_missing_sheet_raises():
    wb = Workbook()
    ws = wb.active
    ws.title = "其他"
    ws.append(["a", "b"])
    bio = io.BytesIO()
    wb.save(bio)
    with pytest.raises(ValueError, match="股票.*not found"):
        parse_capital_xlsx(bio.getvalue())


def test_parse_capital_xlsx_missing_header_raises():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("股票")
    # Wrong column names
    ws.append(["代號", "名稱", "權重", "張數"])
    ws.append(["2330", "台積電", "48.5%", "100,000"])
    bio = io.BytesIO()
    wb.save(bio)
    with pytest.raises(ValueError, match="header row missing"):
        parse_capital_xlsx(bio.getvalue())


def test_parse_capital_xlsx_stops_at_blank_row():
    """Trailing blank rows shouldn't crash; data ends at first blank stock_id."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("股票")
    ws.append(["股票代號", "股票名稱", "持股權重(%)", "股數"])
    ws.append(["2330", "台積電", "48.5%", "100,000"])
    ws.append([None, None, None, None])  # blank row
    ws.append(["XXXX", "後段", "1%", "1"])  # should be ignored after blank
    bio = io.BytesIO()
    wb.save(bio)
    holdings = parse_capital_xlsx(bio.getvalue())
    assert len(holdings) == 1
    assert holdings[0].stock_id == "2330"


# --- manual xlsx file picker ---

def test_find_latest_manual_xlsx_picks_newest_date(tmp_path):
    (tmp_path / "capital_00982A_20260420.xlsx").write_bytes(b"x")
    (tmp_path / "capital_00982A_20260427.xlsx").write_bytes(b"x")
    (tmp_path / "capital_00982A_20260415.xlsx").write_bytes(b"x")
    found = find_latest_manual_xlsx("00982A", tmp_path)
    assert found.name == "capital_00982A_20260427.xlsx"


def test_find_latest_manual_xlsx_ignores_other_tickers(tmp_path):
    (tmp_path / "capital_00992A_20260427.xlsx").write_bytes(b"x")
    assert find_latest_manual_xlsx("00982A", tmp_path) is None


def test_find_latest_manual_xlsx_ignores_malformed_filename(tmp_path):
    (tmp_path / "capital_00982A.xlsx").write_bytes(b"x")  # no date
    (tmp_path / "capital_00982A_April.xlsx").write_bytes(b"x")  # bad date
    assert find_latest_manual_xlsx("00982A", tmp_path) is None


def test_find_latest_manual_xlsx_missing_dir_returns_none(tmp_path):
    assert find_latest_manual_xlsx("00982A", tmp_path / "does-not-exist") is None


# --- SSR meta parser ---

@pytest.mark.parametrize("ticker", ["00982A", "00992A"])
def test_parse_capital_meta_ssr(ticker, fixture_path):
    """SSR portfolio page surfaces nav_total / p_unit / units_outstanding /
    as_of_date even though it only renders top-10 holdings."""
    text = fixture_path(f"capital_{ticker}.html").read_text(encoding="utf-8")
    meta = parse_capital_meta_ssr(text)

    assert meta.get("as_of_date") == "2026-04-27"  # date-picker default at fixture-snapshot time
    assert isinstance(meta.get("nav_total"), float) and meta["nav_total"] > 1e9, \
        "nav_total should be in billions+ TWD, not a stray ngcontent number"
    assert isinstance(meta.get("p_unit"), float) and 0 < meta["p_unit"] < 1000
    assert isinstance(meta.get("units_outstanding"), float) and meta["units_outstanding"] > 1e6


def test_parse_capital_meta_ssr_no_table_returns_empty():
    assert parse_capital_meta_ssr("<html>no info table</html>") == {}
