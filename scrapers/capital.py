"""Capital / 群益投信 (Capital Investment Trust) holdings scraper.

Supports: 00982A (主動群益台灣強棒), 00992A (主動群益科技創新).

Data flow (in priority order):
  1. data/manual/capital_{ticker}_{YYYYMMDD}.xlsx — user drops the official
     "下載資料" xlsx into here for full holdings (~50 stocks).
  2. SSR HTML at /etf/product/detail/{internal_id}/portfolio — fallback when
     no manual xlsx exists; only the top-10 holdings are rendered.

The reason for the manual fallback: the full-holdings JSON lives behind an
internal-network API (125.227.3.107) and the public CDN's mapping pointers
are months stale. See docs/known-issues/capital-scraper.md for the discovery
trail and rationale.
"""
from __future__ import annotations
import io
import logging
import re
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from scrapers.base import BaseScraper, Holding, ScrapeResult, classify_market

logger = logging.getLogger(__name__)

PORTFOLIO_URL = "https://www.capitalfund.com.tw/etf/product/detail/{id}/portfolio"

# Stable ticker → internal product detail id (extracted from main.js routes)
_TICKER_TO_ID = {
    "00982A": 399,
    "00992A": 500,
}

MANUAL_XLSX_DIR = Path("data/manual")

# Excel column headers used by Capital's client-side SheetJS export.
# See exportEtfPortfolio() in their main.js — the "股票" sheet uses these
# exact Chinese keys via json_to_sheet().
_XLSX_SHEET_NAME = "股票"
_XLSX_HDR_STOCK_ID = "股票代號"
_XLSX_HDR_STOCK_NAME = "股票名稱"
_XLSX_HDR_WEIGHT = "持股權重(%)"
_XLSX_HDR_SHARES = "股數"

# Filename pattern: capital_{ticker}_{YYYYMMDD}.xlsx (date sortable lexically)
_MANUAL_FILENAME_RE = re.compile(r"^capital_(?P<ticker>[\w]+)_(?P<date>\d{8})\.xlsx$")


def parse_capital_holdings(text: str) -> list[Holding]:
    """Parse top-10 holdings from a Capital ETF portfolio page (Angular SSR).

    Each holding row is rendered as:
        <div class="tr show-for-medium">
          <div class="th"> {stock_id} </div>
          <div class="th"> {stock_name} </div>
          <div class="td"> {weight}% </div>
          <div class="td sm-full"> {shares with commas} </div>
        </div>

    Fail loudly if the structure no longer matches.
    """
    soup = BeautifulSoup(text, "html.parser")
    rows = soup.find_all(
        "div",
        class_=lambda c: c and "tr" in c.split() and "show-for-medium" in c.split(),
    )
    if not rows:
        raise ValueError(
            "Capital parser: no rows with class 'tr show-for-medium' found — "
            "page structure may have changed"
        )

    holdings: list[Holding] = []
    for row in rows:
        cells = [c.get_text(strip=True) for c in row.find_all("div", recursive=False)]
        if len(cells) != 4:
            raise ValueError(
                f"Capital parser: expected 4 cells per row, got {len(cells)}: {cells!r} — "
                "page structure may have changed"
            )
        stock_id, stock_name, weight_str, shares_str = cells
        if not stock_id or not stock_name:
            raise ValueError(
                f"Capital parser: empty stock_id or name in row {cells!r} — "
                "page structure may have changed"
            )

        try:
            weight_pct = float(weight_str.rstrip("%"))
        except ValueError as exc:
            raise ValueError(
                f"Capital parser: cannot parse weight {weight_str!r} for {stock_id} — "
                "page structure may have changed"
            ) from exc

        try:
            shares = int(shares_str.replace(",", ""))
        except ValueError as exc:
            raise ValueError(
                f"Capital parser: cannot parse shares {shares_str!r} for {stock_id} — "
                "page structure may have changed"
            ) from exc

        holdings.append(
            Holding(
                stock_id=stock_id,
                stock_name=stock_name,
                weight_pct=weight_pct,
                shares=shares,
                market=classify_market(stock_id),
            )
        )

    return holdings


# SSR 表格的 metadata 欄位 — 標籤對應 fund_meta key
_SSR_LABEL_TO_KEY = {
    "基金淨資產價值":   "nav_total",
    "每受益權單位淨資產價值": "p_unit",
    "已發行受益權單位總數":   "units_outstanding",
}
# 抓 td 值欄位內的數字（可能有也可能沒有 <span>TWD</span> 前綴）
# 必須 anchor 在 'td cell auto' 防止抓到 ngcontent attribute 裡的數字
_SSR_VALUE_CELL_RE = re.compile(
    r'class="td cell auto"[^>]*>'
    r'(?:<span[^>]*>TWD</span>\s*)?\s*'
    r'([\d,]+(?:\.\d+)?)'
)
# 資料日期由 datepicker 的預設 value 透露（input-pad-trend 是該欄獨有 class）
_SSR_DATE_RE = re.compile(r'class="[^"]*input-pad-trend[^"]*"\s+value="(\d{4})/(\d{2})/(\d{2})"')


def parse_capital_meta_ssr(text: str) -> dict:
    """Extract fund-level metadata from the SSR portfolio page.

    The Angular Universal HTML SSRs an info table with three labelled rows
    (基金淨資產價值 / 每受益權單位淨資產價值 / 已發行受益權單位總數) and a
    date-picker whose default value is the latest data date.

    Limitations: this is the same SSR that only ships top-10 holdings —
    everything else lives behind the inaccessible internal API. Manual xlsx
    upload (the alternate fetch path) doesn't yet plumb meta through.
    See docs/known-issues/capital-scraper.md.

    Returns {} if the SSR layout changed beyond recognition.
    """
    meta: dict = {}

    for label, key in _SSR_LABEL_TO_KEY.items():
        idx = text.find(label)
        if idx < 0:
            continue
        # Search forward for the next 'td cell auto' value cell — the structure
        # is always label-row </div> then value-row <div class="td cell auto">.
        window = text[idx : idx + 400]
        m = _SSR_VALUE_CELL_RE.search(window)
        if m:
            try:
                meta[key] = float(m.group(1).replace(",", ""))
            except ValueError:
                pass

    m = _SSR_DATE_RE.search(text)
    if m:
        meta["as_of_date"] = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    return meta


def find_latest_manual_xlsx(ticker: str, manual_dir: Path = MANUAL_XLSX_DIR) -> Path | None:
    """Most recent capital_{ticker}_{YYYYMMDD}.xlsx in manual_dir, or None.

    Date-stamped filenames sort lexically; the largest is the freshest.
    Files that don't match the convention are ignored — keeps a stray rename
    from silently overriding the picker.
    """
    if not manual_dir.exists():
        return None
    candidates = []
    for p in manual_dir.glob(f"capital_{ticker}_*.xlsx"):
        m = _MANUAL_FILENAME_RE.match(p.name)
        if m and m.group("ticker") == ticker:
            candidates.append((m.group("date"), p))
    if not candidates:
        return None
    candidates.sort()  # lex sort on YYYYMMDD = chronological
    return candidates[-1][1]


def parse_capital_xlsx(content: bytes) -> list[Holding]:
    """Parse the "股票" sheet from a Capital "下載資料" xlsx.

    Header row uses 股票代號 / 股票名稱 / 持股權重(%) / 股數.
    Weight is stored as the literal string "8.71%"; shares as "1,548,000".
    Fail loudly on structural mismatch.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Capital xlsx parser: cannot open xlsx — file may be corrupt or not actually xlsx"
        ) from exc

    if _XLSX_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Capital xlsx parser: sheet '{_XLSX_SHEET_NAME}' not found "
            f"(sheets: {wb.sheetnames}) — file structure may have changed"
        )
    ws = wb[_XLSX_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Capital xlsx parser: 股票 sheet is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    required = (_XLSX_HDR_STOCK_ID, _XLSX_HDR_STOCK_NAME, _XLSX_HDR_WEIGHT, _XLSX_HDR_SHARES)
    try:
        col = {h: header.index(h) for h in required}
    except ValueError as exc:
        raise ValueError(
            f"Capital xlsx parser: header row missing one of {required} (got {header}) — "
            "file structure may have changed"
        ) from exc

    holdings: list[Holding] = []
    for raw in rows[1:]:
        sid = raw[col[_XLSX_HDR_STOCK_ID]] if col[_XLSX_HDR_STOCK_ID] < len(raw) else None
        if sid is None or str(sid).strip() == "":
            break  # blank row = end of data
        stock_id = str(sid).strip()

        name_raw = raw[col[_XLSX_HDR_STOCK_NAME]]
        if name_raw is None or str(name_raw).strip() == "":
            raise ValueError(
                f"Capital xlsx parser: empty stock_name for {stock_id} — "
                "file structure may have changed"
            )
        stock_name = str(name_raw).strip()

        weight_raw = raw[col[_XLSX_HDR_WEIGHT]]
        try:
            weight_pct = float(str(weight_raw).rstrip("%").strip())
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Capital xlsx parser: cannot parse weight {weight_raw!r} for {stock_id} — "
                "file structure may have changed"
            ) from exc

        shares_raw = raw[col[_XLSX_HDR_SHARES]]
        try:
            shares = int(float(str(shares_raw).replace(",", "")))
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Capital xlsx parser: cannot parse shares {shares_raw!r} for {stock_id} — "
                "file structure may have changed"
            ) from exc

        holdings.append(Holding(
            stock_id=stock_id,
            stock_name=stock_name,
            weight_pct=weight_pct,
            shares=shares,
            market=classify_market(stock_id),
        ))

    if not holdings:
        raise ValueError(
            "Capital xlsx parser: header row found but no data rows extracted — "
            "file structure may have changed"
        )
    return holdings


class CapitalScraper(BaseScraper):
    """Scraper for 群益投信 active ETFs (00982A, 00992A).

    Prefers user-supplied manual xlsx (full ~50 stocks); falls back to SSR
    top-10 when no manual file is present. See docs/known-issues/capital-scraper.md.
    """

    def fetch(self, ticker: str) -> ScrapeResult:
        if ticker not in _TICKER_TO_ID:
            raise ValueError(
                f"CapitalScraper supports {sorted(_TICKER_TO_ID)}, got {ticker!r}"
            )

        manual = find_latest_manual_xlsx(ticker)
        if manual is not None:
            logger.info("capital %s: using manual xlsx %s", ticker, manual.name)
            # TODO: capital's xlsx has an 投資組合 sheet with PCF metadata;
            # add parser when we have a real-world xlsx fixture to base it on.
            return ScrapeResult(
                holdings=parse_capital_xlsx(manual.read_bytes()),
                fund_meta={},
            )

        logger.warning(
            "capital %s: top-10 only (no data/manual/capital_%s_*.xlsx found), "
            "see docs/known-issues/capital-scraper.md",
            ticker, ticker,
        )
        url = PORTFOLIO_URL.format(id=_TICKER_TO_ID[ticker])
        text = self.get(url)
        return ScrapeResult(
            holdings=parse_capital_holdings(text),
            fund_meta=parse_capital_meta_ssr(text),
        )
