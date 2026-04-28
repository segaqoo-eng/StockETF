"""Fuh Hwa / 復華投信 holdings scraper.

Supports: 00991A (主動復華未來50).

Strategy: 2-step fetch.
1. GET the SSR detail page; the "檔案下載" link encodes the latest data date
   as `/api/assetsExcel/{ETF_ID}/{YYYYMMDD}` — Fuh Hwa is the source of truth
   for "what's the most recent trading day with data".
2. GET that xlsx and parse Sheet1 with openpyxl. The header row uses
   證券代號 / 證券名稱 / 股數 / 金額 / 權重(%); holdings begin one row below
   and run until the first empty stock_id cell.

Stock names in the xlsx are 4-char abbreviations (e.g. 台灣積體 for 台積電) —
that's how Fuh Hwa stores them; the cross-aggregation joins by stock_id so
display naming is intentionally left as-is.
"""
from __future__ import annotations
import io
import json
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from scrapers.base import BaseScraper, Holding, ScrapeResult, classify_market

DETAIL_URL = "https://www.fhtrust.com.tw/ETF/etf_detail/{id}"
EXCEL_URL  = "https://www.fhtrust.com.tw/api/assetsExcel/{id}/{date}"
API_URL    = "https://www.fhtrust.com.tw/api/assets?fundID={id}&qDate={date}"

TAIPEI = ZoneInfo("Asia/Taipei")

# Stable ticker → internal product detail id (extracted from page URL on issuer site)
_TICKER_TO_ID = {
    "00991A": "ETF23",
}

_XLSX_HDR_STOCK_ID = "證券代號"  # alias used by the meta parser to know where data starts

# Excel header row identifies the column layout
_HDR_STOCK_ID = "證券代號"
_HDR_STOCK_NAME = "證券名稱"
_HDR_SHARES = "股數"
_HDR_WEIGHT = "權重(%)"

# Pre-header rows carry fund-level metadata as label/value pairs.
# Layout (0-indexed): row 2 holds "日期: YYYY/MM/DD"; rows 3/4, 5/6, 7/8 are
# label-then-value for nav / units / p_unit. We match by label substring so
# row insertions or order shuffles in future xlsx generations don't break us.
_META_LABELS = {
    "nav_total":         "基金資產淨值",
    "units_outstanding": "基金在外流通單位數",
    "p_unit":            "基金每單位淨值",
}
_DATE_LABEL = "日期"
_DATE_RE = re.compile(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})")


def extract_date_from_detail(html: str, internal_id: str) -> str:
    """Pull the YYYYMMDD date from the page's xlsx download link.

    Fail loudly if the link isn't present — the page layout changed.
    """
    pat = r"/api/assetsExcel/" + re.escape(internal_id) + r"/(\d{8})"
    m = re.search(pat, html)
    if not m:
        raise ValueError(
            f"Fuh Hwa parser: assetsExcel link for {internal_id} not found in detail page — "
            "page structure may have changed"
        )
    return m.group(1)


def parse_fuhhwa_meta(content: bytes) -> dict:
    """Extract fund-level metadata from the pre-header rows of the xlsx.

    Returns {} on any parse failure — silent fallback so a metadata regression
    doesn't take down the holdings scrape.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return {}
    if not wb.sheetnames:
        return {}
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    meta: dict = {}

    for i, row in enumerate(rows):
        cell0 = row[0] if row and row[0] is not None else ""
        text = str(cell0).strip()
        if not text:
            continue

        # Date row: "日期: 2026/04/24"
        if _DATE_LABEL in text:
            m = _DATE_RE.search(text)
            if m:
                meta["as_of_date"] = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

        # Label rows precede their value row by 1
        for key, label in _META_LABELS.items():
            if text == label and i + 1 < len(rows):
                next_row = rows[i + 1]
                if next_row and next_row[0] is not None:
                    val = _safe_float(next_row[0])
                    if val is not None:
                        meta[key] = val

        # Stop at the holdings header — no metadata after this point
        if text == _XLSX_HDR_STOCK_ID:
            break

    return meta


def _safe_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def parse_fuhhwa_xlsx(content: bytes) -> list[Holding]:
    """Parse holdings from the Fuh Hwa xlsx download.

    Locates the header row (證券代號/證券名稱/股數/金額/權重(%)) then reads each
    following row until the first row whose stock_id is empty. Fail loudly on
    structural mismatch.
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Fuh Hwa parser: cannot open xlsx — file may be corrupt or not actually xlsx"
        ) from exc

    if not wb.sheetnames:
        raise ValueError("Fuh Hwa parser: workbook has no sheets")
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))

    # Locate header row
    header_idx = None
    col_idx: dict[str, int] = {}
    for i, row in enumerate(rows):
        if row and row[0] == _HDR_STOCK_ID:
            for j, cell in enumerate(row):
                if cell:
                    col_idx[str(cell).strip()] = j
            if all(h in col_idx for h in (_HDR_STOCK_ID, _HDR_STOCK_NAME, _HDR_SHARES, _HDR_WEIGHT)):
                header_idx = i
                break
    if header_idx is None:
        raise ValueError(
            f"Fuh Hwa parser: header row with columns "
            f"{[_HDR_STOCK_ID, _HDR_STOCK_NAME, _HDR_SHARES, _HDR_WEIGHT]} not found — "
            "page structure may have changed"
        )

    holdings: list[Holding] = []
    for row in rows[header_idx + 1 :]:
        raw_id = row[col_idx[_HDR_STOCK_ID]] if col_idx[_HDR_STOCK_ID] < len(row) else None
        if raw_id is None or str(raw_id).strip() == "":
            break
        stock_id = str(raw_id).strip()

        raw_name = row[col_idx[_HDR_STOCK_NAME]] if col_idx[_HDR_STOCK_NAME] < len(row) else None
        if raw_name is None or str(raw_name).strip() == "":
            raise ValueError(
                f"Fuh Hwa parser: empty stock_name for {stock_id} — "
                "page structure may have changed"
            )
        stock_name = str(raw_name).strip()

        raw_shares = row[col_idx[_HDR_SHARES]]
        try:
            # Excel may store as int or as comma-formatted string
            shares = int(float(str(raw_shares).replace(",", "")))
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Fuh Hwa parser: cannot parse shares {raw_shares!r} for {stock_id} — "
                "page structure may have changed"
            ) from exc

        raw_weight = row[col_idx[_HDR_WEIGHT]]
        try:
            # Weight comes as "14.929%" string; strip % and convert
            weight_pct = float(str(raw_weight).rstrip("%").strip())
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Fuh Hwa parser: cannot parse weight {raw_weight!r} for {stock_id} — "
                "page structure may have changed"
            ) from exc

        holdings.append(
            Holding(
                stock_id=stock_id,
                stock_name=stock_name,
                weight_pct=weight_pct,
                shares=shares,
                market=classify_market(stock_id),
            )
        )

    if not holdings:
        raise ValueError(
            "Fuh Hwa parser: header row found but no data rows extracted — "
            "page structure may have changed"
        )
    return holdings


def parse_fuhhwa_api(text: str) -> tuple[list[Holding], dict]:
    """Parse holdings + fund_meta from the /api/assets JSON endpoint.

    Returns ([], {}) on any failure — callers fall back to XLSX.
    """
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return [], {}
    if data.get("status") != 0 or not data.get("result"):
        return [], {}

    fund = data["result"][0]

    # Fund metadata
    meta: dict = {}
    raw_date = fund.get("dDate", "")
    if raw_date:
        meta["as_of_date"] = raw_date.replace("/", "-")
    for key, src in [
        ("nav_total",         "pcf_FundNav"),
        ("units_outstanding", "pcf_FundQissue"),
        ("p_unit",            "pcf_Fundpnav"),
    ]:
        val = _safe_float(str(fund.get(src, "")).replace(",", ""))
        if val is not None:
            meta[key] = val

    # Holdings — skip non-equity rows (ftype != "股票")
    holdings: list[Holding] = []
    for item in fund.get("detail", []):
        if item.get("ftype") != "股票":
            continue
        stock_id = str(item.get("stockid", "")).strip()
        if not stock_id:
            continue
        stock_name = str(item.get("stockname", "")).strip()
        try:
            shares = int(float(str(item.get("qshare", "0")).replace(",", "")))
        except (TypeError, ValueError):
            shares = 0
        try:
            weight_pct = float(str(item.get("prate_addaccint", "0")).rstrip("%").strip())
        except (TypeError, ValueError):
            weight_pct = 0.0
        holdings.append(Holding(
            stock_id=stock_id,
            stock_name=stock_name,
            weight_pct=weight_pct,
            shares=shares,
            market=classify_market(stock_id),
        ))

    return holdings, meta


class FuhhwaScraper(BaseScraper):
    """Scraper for 復華投信 active ETFs (00991A).

    2-step fetch: detail page (date) → xlsx (holdings).
    """

    def fetch(self, ticker: str) -> ScrapeResult:
        if ticker not in _TICKER_TO_ID:
            raise ValueError(
                f"FuhhwaScraper supports {sorted(_TICKER_TO_ID)}, got {ticker!r}"
            )
        internal_id = _TICKER_TO_ID[ticker]
        today = datetime.now(TAIPEI).strftime("%Y/%m/%d")

        # Primary: JSON API (today's date, more current than XLSX)
        try:
            api_text = self.get(API_URL.format(id=internal_id, date=today))
            holdings, fund_meta = parse_fuhhwa_api(api_text)
            if holdings:
                return ScrapeResult(holdings=holdings, fund_meta=fund_meta)
        except Exception:
            pass

        # Fallback: dated XLSX (same as before)
        detail_html = self.get(DETAIL_URL.format(id=internal_id))
        date = extract_date_from_detail(detail_html, internal_id)
        xlsx_bytes = self.get_bytes(EXCEL_URL.format(id=internal_id, date=date))
        return ScrapeResult(
            holdings=parse_fuhhwa_xlsx(xlsx_bytes),
            fund_meta=parse_fuhhwa_meta(xlsx_bytes),
        )
